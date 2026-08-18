import uuid
from datetime import datetime
from typing import Optional
from fastapi import HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_

from app.models.application import Application, ApplicationStageHistory
from app.schemas.application import (
    ApplicationCreate, ApplicationResponse, ApplicationDetailResponse,
    ApplicantBrief, StageHistoryEntry,
)
from app.constants.stages import VALID_TRANSITIONS, STAGE_LABELS, REASON_REQUIRED_STAGES, MOVE_JOB_ALLOWED_STAGES


async def _seed_initial_resume_version(db: AsyncSession, application: Application, applicant_id: uuid.UUID) -> None:
    """Give a brand-new application its v1 resume-history row immediately.

    The Resume tab (`ResumeVersions.jsx`) only ever reads `application_resumes`,
    never `Application.resume_url` directly — so without this, every application
    submitted after the one-time f7a8b9c0d1e2 backfill migration shows "no resume
    on file" despite `resume_url` being set. Mirrors that migration's own
    INSERT exactly (uploaded_by/role = the applicant), so the version history
    reads the same regardless of whether v1 came from the backfill or from here.
    Best-effort: a failure here must never block the application submission
    itself, the same way the notification email sends around it are wrapped.
    """
    if not application.resume_url:
        return
    try:
        from app.models.application_resume import ApplicationResume
        db.add(ApplicationResume(
            application_id=application.id,
            version=1,
            file_url=application.resume_url,
            uploaded_by=applicant_id,
            uploaded_by_role="applicant",
        ))
        await db.commit()
    except Exception:
        await db.rollback()


def _normalize_url(url: Optional[str]) -> Optional[str]:
    """Candidate-supplied LinkedIn/GitHub/portfolio links come in through
    several paths (AI resume parsing, manual HR/agency intake, bulk Excel
    import) and not all of them guarantee a scheme — a bare "linkedin.com/in/x"
    renders as a valid-looking <a href> that the browser then resolves
    *relative to this site's own origin*, so clicking it just reloads the
    careers portal instead of navigating to LinkedIn. Normalizing once here,
    at the single read path every application response goes through, fixes
    it regardless of which ingestion path let the bare value through —
    including rows already stored that way before this fix existed."""
    if not url:
        return url
    url = url.strip()
    return url if url.startswith(("http://", "https://")) else f"https://{url}"


def _app_to_dict(app: Application) -> dict:
    from app.services.storage_service import refresh_url

    return {
        "id": app.id,
        "job_id": app.job_id,
        "applicant_id": app.applicant_id,
        "referral_id": app.referral_id,
        # Re-signed on every read — the SAS token baked in at upload time only
        # lasts 7 days, so serving the stored value as-is breaks any resume
        # older than that with an "AuthenticationFailed ... signed expiry"
        # error from Azure.
        "resume_url": refresh_url(app.resume_url),
        "cover_letter": app.cover_letter,
        "linkedin_url": _normalize_url(app.linkedin_url),
        "portfolio_url": _normalize_url(app.portfolio_url),
        "github_url": _normalize_url(app.github_url),
        "current_ctc": app.current_ctc,
        "expected_ctc": app.expected_ctc,
        "notice_period": app.notice_period,
        "answers": app.answers,
        "stage": app.stage,
        "rejection_reason": app.rejection_reason,
        "drop_category": app.drop_category,
        "source": app.source,
        "agency_id": app.agency_id,
        "rating": app.rating,
        "is_starred": app.is_starred,
        "assigned_to": app.assigned_to,
        "on_hold": app.on_hold,
        "hold_reason": app.hold_reason,
        "duplicate_flag": app.duplicate_flag,
        "duplicate_reason": app.duplicate_reason,
        "duplicate_reviewed_at": app.duplicate_reviewed_at,
        "duplicate_reviewed_by": app.duplicate_reviewed_by,
        "applied_at": app.applied_at,
        "stage_updated_at": app.stage_updated_at,
        "created_at": app.created_at,
        "updated_at": app.updated_at,
    }

FREE_TEXT_MAX = 255

_FREE_TEXT_LABELS = {
    "current_ctc": "Current CTC",
    "expected_ctc": "Expected CTC",
    "notice_period": "Notice period",
}


def _validate_free_text_lengths(**values: Optional[str]) -> None:
    too_long = [
        f"{_FREE_TEXT_LABELS[field]} ({len(value.strip())} characters)"
        for field, value in values.items()
        if value and len(value.strip()) > FREE_TEXT_MAX
    ]
    if too_long:
        raise HTTPException(
            400,
            f"Please shorten to under {FREE_TEXT_MAX} characters: {', '.join(too_long)}",
        )


async def submit_application(
    db: AsyncSession,
    data: ApplicationCreate,
    applicant_id: uuid.UUID,
) -> Application:
    from app.services import duplicate_detection_service as dupes

    _validate_free_text_lengths(
        current_ctc=data.current_ctc,
        expected_ctc=data.expected_ctc,
        notice_period=data.notice_period,
    )

    # Block reapplication within 6 months of any rejection — same account, so this
    # is authoritative (unlike the fuzzy name match below).
    recent_rejection = await dupes.check_cooloff(db, applicant_id)
    if recent_rejection:
        raise HTTPException(403, dupes.cooloff_message(recent_rejection))

    existing = await db.execute(
        select(Application).where(
            Application.job_id == data.job_id,
            Application.applicant_id == applicant_id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(409, "You have already applied for this job")

    agency_id = None
    agency_ref = data.agency_ref
    if agency_ref:
        from app.models.agency import JobAgencyAssignment
        from app.models.application import Application as App
        assignment = (await db.execute(
            select(JobAgencyAssignment).where(JobAgencyAssignment.ref_token == agency_ref)
        )).scalar_one_or_none()
        if assignment and str(assignment.job_id) == str(data.job_id):
            agency_id = assignment.agency_id

    # Persist candidate profile fields (profile is the source of truth)
    from app.schemas.application import PROFILE_FIELDS
    from app.models.candidate_profile import CandidateProfile
    from app.models.user import User as UserModel

    user = await db.get(UserModel, applicant_id)
    if user and data.date_of_birth:
        user.date_of_birth = data.date_of_birth

    # Validate the referral (if any) rather than trusting a raw UUID from the
    # client — silently ignore a stale/expired/email-mismatched referral
    # instead of blocking the apply, consistent with how a mismatched
    # agency_ref is handled above (best-effort attribution, not a hard gate).
    referral = None
    if data.referral_id:
        from datetime import timezone as _tz
        from app.models.referral import Referral
        candidate = await db.get(Referral, data.referral_id)
        if candidate:
            now = datetime.now(_tz.utc)
            expired = candidate.status == "expired" or (candidate.expires_at and candidate.expires_at < now)
            email_matches = user and candidate.candidate_email.lower() == (user.email or "").lower()
            # A referral is scoped to the one job it was created for — without
            # this, a valid (unexpired, matching-email) referral id could be
            # replayed on a *different* job's apply form and still get
            # attributed as a referral there, letting one invite link silently
            # earn referral credit across jobs it was never actually raised for.
            job_matches = str(candidate.job_id) == str(data.job_id)
            if not expired and email_matches and job_matches:
                referral = candidate

    profile = await db.get(CandidateProfile, applicant_id)
    if not profile:
        profile = CandidateProfile(user_id=applicant_id)
        db.add(profile)
    profile.current_company = data.current_company
    profile.current_designation = data.current_designation
    profile.total_experience = data.total_experience
    profile.current_location = data.current_location
    profile.education = data.education
    if data.skills is not None:
        profile.skills = data.skills

    duplicate_flag, duplicate_reason = (False, None)
    if user and user.full_name:
        duplicate_flag, duplicate_reason = await dupes.build_duplicate_flag(
            db, user=user, full_name=user.full_name,
        )

    source = "referral" if referral else ("agency" if agency_id else "direct")

    # Job-level visibility gate: an internal-only job has no external apply
    # route at all; a non-internal job still needs its matching flag on for
    # whichever route this submission is coming through (agency submissions
    # go through a separate authenticated portal flow and aren't gated here).
    from app.models.job import Job as JobModel
    job = await db.get(JobModel, data.job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    if source == "referral" and (job.is_internal or not job.allow_referrals):
        raise HTTPException(403, "This job is not open to referral applications")
    if source == "direct" and (job.is_internal or not job.allow_outsiders):
        raise HTTPException(403, "This job is not open to public applications")

    create_data = data.model_dump(exclude={"agency_ref", "referral_id", *PROFILE_FIELDS})
    application = Application(
        applicant_id=applicant_id,
        source=source,
        agency_id=agency_id,
        referral_id=referral.id if referral else None,
        duplicate_flag=duplicate_flag,
        duplicate_reason=duplicate_reason,
        **create_data,
    )
    db.add(application)
    await db.commit()
    await db.refresh(application)

    await _seed_initial_resume_version(db, application, applicant_id)

    if referral and referral.status in ("pending", "invited"):
        try:
            from app.services import referral_service
            await referral_service.update_status(db, referral.id, "applied")
        except Exception:
            pass

    try:
        from app.tasks.email_tasks import send_application_received_email
        send_application_received_email.delay(str(application.id))
    except Exception:
        pass

    if job.screening_enabled:
        try:
            from app.services import screening_service
            await screening_service.create_and_queue_email(db, application.id)
        except Exception:
            pass

    return application


async def submit_sourced_application(
    db: AsyncSession,
    *,
    job_id: uuid.UUID,
    full_name: str,
    email: str,
    resume_url: str,
    source: str,
    agency_id: Optional[uuid.UUID] = None,
    phone: Optional[str] = None,
    linkedin_url: Optional[str] = None,
    portfolio_url: Optional[str] = None,
    github_url: Optional[str] = None,
    cover_letter: Optional[str] = None,
    current_location: Optional[str] = None,
    total_experience: Optional[str] = None,
    current_company: Optional[str] = None,
    current_designation: Optional[str] = None,
    education: Optional[str] = None,
    skills: Optional[str] = None,
    current_ctc: Optional[str] = None,
    expected_ctc: Optional[str] = None,
    notice_period: Optional[str] = None,
) -> Application:
    """Create an application on behalf of a candidate (agency upload or HR/TA sourcing).

    Finds the candidate user by email or creates a passwordless account (they can
    claim it via the forgot-password flow). The 6-month rejection cool-off applies
    here exactly as it does to a direct candidate apply — the source of the
    submission doesn't change who the candidate is, so it can't be used to route
    around the reapply block."""
    import secrets
    from datetime import timedelta, timezone
    from app.models.user import User as UserModel
    from app.models.candidate_profile import CandidateProfile
    from app.models.job import Job
    from app.utils.security import hash_password, generate_token
    from app.services import duplicate_detection_service as dupes

    # Before any writes — a failure here must not leave a half-created user.
    _validate_free_text_lengths(
        current_ctc=current_ctc,
        expected_ctc=expected_ctc,
        notice_period=notice_period,
    )

    job = await db.get(Job, job_id)
    if not job:
        raise HTTPException(404, "Job not found")

    email = email.strip().lower()
    user = (await db.execute(
        select(UserModel).where(func.lower(UserModel.email) == email)
    )).scalar_one_or_none()

    if user and user.role not in ("applicant", "employee"):
        raise HTTPException(400, "This email belongs to an internal user account")

    is_new_user = not user
    if not user:
        # No password is set on a sourced account — the candidate claims it via the
        # set-password link in send_sourced_application_welcome_email (a proactively
        # generated password_reset_token), or the standard /forgot-password flow later.
        user = UserModel(
            email=email,
            full_name=full_name.strip(),
            password_hash=hash_password(secrets.token_urlsafe(24)),
            role="applicant",
            phone=phone,
            is_verified=False,
            password_reset_token=generate_token(),
            password_reset_expires=datetime.now(timezone.utc) + timedelta(days=7),
        )
        db.add(user)
        await db.flush()
    elif phone and not user.phone:
        user.phone = phone

    existing = (await db.execute(
        select(Application).where(
            Application.job_id == job_id,
            Application.applicant_id == user.id,
        )
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(409, "This candidate has already been submitted for this job")

    if not is_new_user:
        recent_rejection = await dupes.check_cooloff(db, user.id)
        if recent_rejection:
            raise HTTPException(403, dupes.cooloff_message(recent_rejection))

    duplicate_flag, duplicate_reason = await dupes.build_duplicate_flag(
        db, user=user, full_name=full_name,
    )

    profile = await db.get(CandidateProfile, user.id)
    if not profile:
        profile = CandidateProfile(user_id=user.id)
        db.add(profile)
    for field, val in (
        ("current_location", current_location),
        ("total_experience", total_experience),
        ("current_company", current_company),
        ("current_designation", current_designation),
        ("education", education),
        ("skills", skills),
    ):
        if val and val.strip():
            setattr(profile, field, val.strip())

    application = Application(
        job_id=job_id,
        applicant_id=user.id,
        resume_url=resume_url,
        source=source,
        agency_id=agency_id,
        cover_letter=cover_letter,
        linkedin_url=linkedin_url,
        portfolio_url=portfolio_url,
        github_url=github_url,
        current_ctc=(current_ctc.strip() if current_ctc and current_ctc.strip() else None),
        expected_ctc=(expected_ctc.strip() if expected_ctc and expected_ctc.strip() else None),
        notice_period=(notice_period.strip() if notice_period and notice_period.strip() else None),
        duplicate_flag=duplicate_flag,
        duplicate_reason=duplicate_reason,
    )
    db.add(application)
    await db.commit()
    await db.refresh(application)

    await _seed_initial_resume_version(db, application, user.id)

    try:
        if is_new_user:
            from app.tasks.email_tasks import send_sourced_application_welcome_email
            send_sourced_application_welcome_email.delay(str(application.id))
        else:
            from app.tasks.email_tasks import send_application_received_email
            send_application_received_email.delay(str(application.id))
    except Exception:
        pass

    if job.screening_enabled:
        try:
            from app.services import screening_service
            await screening_service.create_and_queue_email(db, application.id)
        except Exception:
            pass

    return application


MAX_BULK_RESUMES = 20
MAX_BULK_EXCEL_ROWS = 300


async def bulk_submit_from_resumes(
    db: AsyncSession,
    *,
    job_id: uuid.UUID,
    source: str,
    files: list[tuple[str, bytes, str]],
    agency_id: Optional[uuid.UUID] = None,
) -> list[dict]:
    """Parse each uploaded resume and create a sourced application from it —
    no manual re-entry per candidate, since that would defeat the point of a
    bulk upload. Each file is fully independent: a failure on one (duplicate
    application, unreadable file, missing name/email) is recorded and the rest
    still get processed. The session is rolled back after every failure so a
    partially-flushed row from one file can never leak into the next file's
    transaction (see note below on why that matters)."""
    from app.services import resume_parsing_service, storage_service

    results: list[dict] = []
    for filename, content, content_type in files:
        try:
            parsed = await resume_parsing_service.parse_resume(content, content_type, filename)
            full_name = parsed.get("full_name")
            email = parsed.get("email")
            if not full_name or not email:
                results.append({
                    "filename": filename,
                    "status": "error",
                    "error": "Could not detect a name and email on this resume — add this candidate manually.",
                })
                continue

            resume_url = await storage_service.upload_resume_bytes(content, filename, content_type, "hr-bulk")
            application = await submit_sourced_application(
                db,
                job_id=job_id,
                full_name=full_name,
                email=email,
                resume_url=resume_url,
                source=source,
                agency_id=agency_id,
                phone=parsed.get("phone"),
                linkedin_url=parsed.get("linkedin_url"),
                portfolio_url=parsed.get("portfolio_url"),
                github_url=parsed.get("github_url"),
                current_location=parsed.get("current_location"),
                total_experience=parsed.get("total_experience"),
                current_company=parsed.get("current_company"),
                current_designation=parsed.get("current_designation"),
                education=parsed.get("education"),
                skills=parsed.get("skills"),
            )
            results.append({
                "filename": filename,
                "status": "success",
                "application_id": str(application.id),
                "candidate_name": full_name,
                "email": email,
                "duplicate_flag": application.duplicate_flag,
                "duplicate_reason": application.duplicate_reason,
            })
        except HTTPException as exc:
            await db.rollback()
            results.append({"filename": filename, "status": "error", "error": str(exc.detail)})
        except Exception:
            await db.rollback()
            results.append({"filename": filename, "status": "error", "error": "Unexpected error while processing this file."})

    return results


# Column headers accepted in the bulk-upload spreadsheet, matched case/space-insensitively.
EXCEL_COLUMN_ALIASES = {
    "full name": "full_name", "name": "full_name", "candidate name": "full_name",
    "email": "email", "email address": "email",
    "phone": "phone", "phone number": "phone", "mobile": "phone",
    "current location": "current_location", "location": "current_location",
    "total experience": "total_experience", "experience": "total_experience",
    "current company": "current_company", "company": "current_company",
    "current designation": "current_designation", "designation": "current_designation", "title": "current_designation",
    "education": "education",
    "skills": "skills",
    "linkedin": "linkedin_url", "linkedin url": "linkedin_url",
    "current ctc": "current_ctc", "current_ctc": "current_ctc",
    "expected ctc": "expected_ctc", "expected_ctc": "expected_ctc",
    "notice period": "notice_period", "notice_period": "notice_period",
}


def parse_bulk_excel(content: bytes) -> list[dict]:
    """Reads the bulk-upload spreadsheet into a list of row dicts keyed by our
    field names (header matching is case/space-insensitive via EXCEL_COLUMN_ALIASES).
    Raises HTTPException(400) if it isn't a readable spreadsheet or has no
    recognizable header row."""
    import io
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
    except Exception as exc:
        raise HTTPException(400, f"Could not read this spreadsheet: {exc}")

    if not header:
        raise HTTPException(400, "The spreadsheet is empty.")

    field_by_col = {}
    for i, cell in enumerate(header):
        key = (str(cell).strip().lower() if cell is not None else "")
        if key in EXCEL_COLUMN_ALIASES:
            field_by_col[i] = EXCEL_COLUMN_ALIASES[key]

    if "full_name" not in field_by_col.values() or "email" not in field_by_col.values():
        raise HTTPException(
            400,
            "The spreadsheet must have at least 'Full Name' and 'Email' columns "
            "(Phone, Current Location, Total Experience, Current Company, "
            "Current Designation, Education, Skills, LinkedIn, Current CTC, "
            "Expected CTC and Notice Period are optional).",
        )

    # Row numbers are tracked against the real spreadsheet (header = row 1) so
    # results reported back to HR line up with what they see in Excel — rows
    # with some data but a missing name/email are kept (not silently dropped)
    # so bulk_submit_from_excel can report them as failed rows instead of the
    # row count just quietly not matching what was uploaded.
    rows = []
    for row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None for c in row):
            continue
        record = {"_row": row_num}
        for i, val in field_by_col.items():
            if i < len(row) and row[i] is not None:
                record[val] = str(row[i]).strip()
        rows.append(record)

    return rows


async def bulk_submit_from_excel(
    db: AsyncSession,
    *,
    job_id: uuid.UUID,
    source: str,
    rows: list[dict],
    agency_id: Optional[uuid.UUID] = None,
) -> list[dict]:
    """Create a sourced application per spreadsheet row. No resume is required
    for this path (resume_url is stored as '' — the Resume tab shows a
    'no resume provided' state for these instead of a broken preview)."""
    results: list[dict] = []
    for row in rows:
        i = row.get("_row")
        full_name = row.get("full_name")
        email = row.get("email")
        if not full_name or not email:
            results.append({
                "row": i, "status": "error", "candidate_name": full_name, "email": email,
                "error": "Missing Full Name or Email — this row was skipped.",
            })
            continue
        try:
            application = await submit_sourced_application(
                db,
                job_id=job_id,
                full_name=full_name,
                email=email,
                resume_url="",
                source=source,
                agency_id=agency_id,
                phone=row.get("phone"),
                linkedin_url=row.get("linkedin_url"),
                current_location=row.get("current_location"),
                total_experience=row.get("total_experience"),
                current_company=row.get("current_company"),
                current_designation=row.get("current_designation"),
                education=row.get("education"),
                skills=row.get("skills"),
                current_ctc=row.get("current_ctc"),
                expected_ctc=row.get("expected_ctc"),
                notice_period=row.get("notice_period"),
            )
            results.append({
                "row": i, "status": "success", "application_id": str(application.id),
                "candidate_name": full_name, "email": email,
                "duplicate_flag": application.duplicate_flag,
                "duplicate_reason": application.duplicate_reason,
            })
        except HTTPException as exc:
            await db.rollback()
            results.append({"row": i, "status": "error", "candidate_name": full_name, "email": email, "error": str(exc.detail)})
        except Exception:
            await db.rollback()
            results.append({"row": i, "status": "error", "candidate_name": full_name, "email": email, "error": "Unexpected error processing this row."})

    return results


async def get_my_applications(
    db: AsyncSession,
    applicant_id: uuid.UUID,
    page: int = 1,
    limit: int = 20,
) -> dict:
    from app.models.job import Job
    from app.models.agency import Agency

    condition = Application.applicant_id == applicant_id
    total = (
        await db.execute(select(func.count()).select_from(Application).where(condition))
    ).scalar_one()

    offset = (page - 1) * limit
    rows = (
        await db.execute(
            select(Application, Job.title.label("job_title"), Agency.name.label("agency_name"))
            .join(Job, Job.id == Application.job_id)
            .join(Agency, Agency.id == Application.agency_id, isouter=True)
            .where(condition)
            .order_by(Application.applied_at.desc())
            .offset(offset)
            .limit(limit)
        )
    ).all()

    items = []
    for app, job_title, agency_name in rows:
        d = _app_to_dict(app)
        d["job_title"] = job_title
        d["agency_name"] = agency_name
        items.append(ApplicationResponse.model_validate(d))

    return {"items": items, "total": total, "page": page, "limit": limit, "pages": max(1, -(-total // limit))}


async def export_applications(
    db: AsyncSession,
    job_id: Optional[uuid.UUID] = None,
    stage: Optional[str] = None,
) -> list[dict]:
    from app.models.user import User
    from app.models.job import Job

    base = (
        select(Application, User.full_name, User.email, Job.title.label("job_title"))
        .join(User, User.id == Application.applicant_id)
        .join(Job, Job.id == Application.job_id)
    )
    filters = []
    if job_id:
        filters.append(Application.job_id == job_id)
    if stage:
        filters.append(Application.stage == stage)
    if filters:
        base = base.where(and_(*filters))

    rows = (await db.execute(base.order_by(Application.applied_at.desc()))).all()
    result = []
    for app, full_name, email, job_title in rows:
        result.append({
            "id": str(app.id),
            "candidate_name": full_name,
            "candidate_email": email,
            "job_title": job_title,
            "stage": app.stage,
            "source": app.source,
            "rating": app.rating or "",
            "applied_at": app.applied_at.strftime("%Y-%m-%d %H:%M") if app.applied_at else "",
            "stage_updated_at": app.stage_updated_at.strftime("%Y-%m-%d %H:%M") if app.stage_updated_at else "",
            "linkedin_url": app.linkedin_url or "",
            "portfolio_url": app.portfolio_url or "",
            "github_url": app.github_url or "",
        })
    return result


async def get_all_applications(
    db: AsyncSession,
    *,
    job_id: Optional[uuid.UUID] = None,
    stage: Optional[str] = None,
    search: Optional[str] = None,
    agency_id: Optional[uuid.UUID] = None,
    source: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
) -> dict:
    from app.models.user import User
    from app.models.agency import Agency
    from app.models.screening import ScreeningResponse

    base = (
        select(
            Application, User.full_name, User.email, User.avatar_url, Agency.name.label("agency_name"),
            ScreeningResponse.overall_score, ScreeningResponse.auto_reject,
        )
        .join(User, User.id == Application.applicant_id)
        .join(Agency, Agency.id == Application.agency_id, isouter=True)
        .join(ScreeningResponse, ScreeningResponse.application_id == Application.id, isouter=True)
    )

    filters = []
    if job_id:
        filters.append(Application.job_id == job_id)
    if stage:
        filters.append(Application.stage == stage)
    if search:
        filters.append(User.full_name.ilike(f"%{search}%"))
    if agency_id:
        filters.append(Application.agency_id == agency_id)
    if source:
        filters.append(Application.source == source)

    if filters:
        base = base.where(and_(*filters))

    count_stmt = (
        select(func.count()).select_from(Application)
        .join(User, User.id == Application.applicant_id)
    )
    if filters:
        count_stmt = count_stmt.where(and_(*filters))
    total = (await db.execute(count_stmt)).scalar_one()

    offset = (page - 1) * limit
    rows = (await db.execute(
        base.order_by(Application.applied_at.desc()).offset(offset).limit(limit)
    )).all()

    items = []
    for app, full_name, email, avatar_url, agency_name, screening_score, screening_auto_reject in rows:
        d = _app_to_dict(app)
        d["agency_name"] = agency_name
        d["screening_score"] = float(screening_score) if screening_score is not None else None
        d["screening_auto_reject"] = screening_auto_reject
        d["applicant"] = {
            "id": app.applicant_id,
            "full_name": full_name,
            "email": email,
            "avatar_url": avatar_url,
        }
        items.append(ApplicationResponse.model_validate(d))

    return {"items": items, "total": total, "page": page, "limit": limit, "pages": max(1, -(-total // limit))}


async def get_application_by_id(db: AsyncSession, application_id: uuid.UUID) -> ApplicationDetailResponse:
    from sqlalchemy.orm import aliased
    from app.models.user import User
    from app.models.interview import Interview
    from app.models.candidate_profile import CandidateProfile
    from app.models.agency import Agency
    from app.models.referral import Referral

    Referrer = aliased(User)

    row = (await db.execute(
        select(
            Application, User.full_name, User.email, User.avatar_url, User.date_of_birth,
            Agency.name.label("agency_name"), Referrer.full_name.label("referrer_name"),
        )
        .join(User, User.id == Application.applicant_id)
        .join(Agency, Agency.id == Application.agency_id, isouter=True)
        .join(Referral, Referral.id == Application.referral_id, isouter=True)
        .join(Referrer, Referrer.id == Referral.referred_by, isouter=True)
        .where(Application.id == application_id)
    )).first()

    if not row:
        raise HTTPException(404, "Application not found")

    app, full_name, email, avatar_url, date_of_birth, agency_name, referrer_name = row
    profile = await db.get(CandidateProfile, app.applicant_id)

    history = (await db.execute(
        select(ApplicationStageHistory)
        .where(ApplicationStageHistory.application_id == application_id)
        .order_by(ApplicationStageHistory.created_at.asc())
    )).scalars().all()

    interview_count = (await db.execute(
        select(func.count()).select_from(Interview).where(Interview.application_id == application_id)
    )).scalar_one()

    d = _app_to_dict(app)
    d["agency_name"] = agency_name
    d["referrer_name"] = referrer_name
    d["applicant"] = {
        "id": app.applicant_id,
        "full_name": full_name,
        "email": email,
        "avatar_url": avatar_url,
    }
    d["stage_history"] = [
        {
            "id": h.id,
            "from_stage": h.from_stage,
            "to_stage": h.to_stage,
            "notes": h.notes,
            "changed_by": h.changed_by,
            "created_at": h.created_at,
        }
        for h in history
    ]
    d["interview_count"] = interview_count
    d["date_of_birth"] = date_of_birth
    d["candidate_profile"] = {
        "current_company": profile.current_company if profile else None,
        "current_designation": profile.current_designation if profile else None,
        "total_experience": profile.total_experience if profile else None,
        "current_location": profile.current_location if profile else None,
        "skills": profile.skills if profile else None,
        "education": profile.education if profile else None,
    }

    return ApplicationDetailResponse.model_validate(d)


async def move_stage(
    db: AsyncSession,
    application_id: uuid.UUID,
    new_stage: str,
    moved_by: uuid.UUID,
    notes: Optional[str] = None,
    rejection_reason: Optional[str] = None,
    drop_category: Optional[str] = None,
) -> Application:
    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")

    if app.on_hold:
        raise HTTPException(400, "Candidate is on hold — resume before changing stage.")

    allowed = VALID_TRANSITIONS.get(app.stage, [])
    if new_stage not in allowed:
        raise HTTPException(400, f"Cannot move from '{app.stage}' to '{new_stage}'")

    if new_stage == "hired":
        from app.models.offer import OfferLetter
        from sqlalchemy import select as _select
        offer = (await db.execute(
            _select(OfferLetter).where(OfferLetter.application_id == application_id)
        )).scalar_one_or_none()
        if not offer or offer.status != "accepted" or not offer.candidate_signature:
            raise HTTPException(
                400,
                "Cannot mark as hired — candidate must accept and digitally sign the offer letter first."
            )

    from_stage = app.stage

    history = ApplicationStageHistory(
        application_id=application_id,
        from_stage=from_stage,
        to_stage=new_stage,
        changed_by=moved_by,
        notes=notes,
    )
    db.add(history)

    app.stage = new_stage
    app.stage_updated_at = datetime.utcnow()
    if new_stage in REASON_REQUIRED_STAGES:
        if rejection_reason:
            app.rejection_reason = rejection_reason
        if drop_category:
            app.drop_category = drop_category

    try:
        from app.models.notification import Notification
        label = STAGE_LABELS.get(new_stage, new_stage.replace("_", " ").title())
        notif = Notification(
            user_id=app.applicant_id,
            type="stage_update",
            title=f"Application update: {label}",
            body="Your application status has been updated. View your applications for details.",
            link="/portal/applications",
        )
        db.add(notif)
    except Exception:
        pass

    await db.commit()
    await db.refresh(app)

    if app.referral_id and new_stage in ("hired", "rejected"):
        try:
            from app.services import referral_service
            await referral_service.update_status(db, app.referral_id, new_stage)
        except Exception:
            pass

    try:
        from app.tasks.email_tasks import send_stage_update_email
        send_stage_update_email.delay(str(application_id), new_stage, from_stage)
    except Exception:
        pass

    if app.agency_id:
        try:
            from app.tasks.email_tasks import send_agency_stage_update_email
            send_agency_stage_update_email.delay(str(application_id), new_stage)
        except Exception:
            pass

    if new_stage == "offer":
        try:
            from app.services.document_service import get_or_create_request
            await get_or_create_request(db, application_id)
            from app.tasks.email_tasks import send_document_request_email_task
            send_document_request_email_task.delay(str(application_id))
        except Exception:
            pass

    return app


async def move_application_job(
    db: AsyncSession,
    application_id: uuid.UUID,
    new_job_id: uuid.UUID,
    moved_by: uuid.UUID,
) -> Application:
    """Reassign an application to a different job req — same application row,
    same resume/history/stage, just pointed at the correct opening (e.g. HR
    finds a screening candidate is a better fit for a sibling role). Kept to
    early stages only: UniqueConstraint(job_id, applicant_id) means the
    candidate can't already have a separate application for the target job,
    and once real interviews are underway they're tied to the original role."""
    from app.models.job import Job as JobModel

    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")

    if app.stage not in MOVE_JOB_ALLOWED_STAGES:
        allowed_labels = " / ".join(STAGE_LABELS.get(s, s) for s in MOVE_JOB_ALLOWED_STAGES)
        raise HTTPException(
            400,
            f"Can only move a candidate to another job while in the {allowed_labels} stage.",
        )

    if new_job_id == app.job_id:
        raise HTTPException(400, "Candidate is already applied to this job")

    new_job = await db.get(JobModel, new_job_id)
    if not new_job:
        raise HTTPException(404, "Target job not found")

    conflict = (await db.execute(
        select(Application).where(
            Application.job_id == new_job_id,
            Application.applicant_id == app.applicant_id,
        )
    )).scalar_one_or_none()
    if conflict:
        raise HTTPException(409, "This candidate already has a separate application for that job")

    old_job = await db.get(JobModel, app.job_id)

    db.add(ApplicationStageHistory(
        application_id=application_id,
        from_stage=app.stage,
        to_stage=app.stage,
        changed_by=moved_by,
        notes=f'Moved from "{old_job.title if old_job else "a previous role"}" to "{new_job.title}"',
    ))

    app.job_id = new_job_id

    try:
        from app.models.notification import Notification
        db.add(Notification(
            user_id=app.applicant_id,
            type="job_moved",
            title="Your application was moved to a different role",
            body=f"Your application has been moved to {new_job.title}.",
            link="/portal/applications",
        ))
    except Exception:
        pass

    await db.commit()
    await db.refresh(app)
    return app


async def dismiss_duplicate_flag(db: AsyncSession, application_id: uuid.UUID, reviewer_id: uuid.UUID) -> Application:
    """HR has looked at the possible-duplicate match and decided this application
    is fine as-is. duplicate_flag itself is left set (audit trail of what was
    found); duplicate_reviewed_at/_by is what the frontend actually checks to
    decide whether to keep showing the review banner."""
    from datetime import timezone

    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")
    app.duplicate_reviewed_at = datetime.now(timezone.utc)
    app.duplicate_reviewed_by = reviewer_id
    await db.commit()
    await db.refresh(app)
    return app


async def toggle_star(db: AsyncSession, application_id: uuid.UUID) -> Application:
    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")
    app.is_starred = not app.is_starred
    await db.commit()
    await db.refresh(app)
    return app


async def set_rating(db: AsyncSession, application_id: uuid.UUID, rating: Optional[int]) -> Application:
    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")
    app.rating = rating
    await db.commit()
    await db.refresh(app)
    return app


async def assign_application(
    db: AsyncSession,
    application_id: uuid.UUID,
    assignee_id: Optional[uuid.UUID],
) -> Application:
    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")
    app.assigned_to = assignee_id
    await db.commit()
    await db.refresh(app)
    return app


async def set_hold(
    db: AsyncSession,
    application_id: uuid.UUID,
    on_hold: bool,
    hold_reason: Optional[str],
) -> Application:
    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")
    app.on_hold = on_hold
    app.hold_reason = hold_reason if on_hold else None
    await db.commit()
    await db.refresh(app)
    return app


async def add_note(
    db: AsyncSession,
    application_id: uuid.UUID,
    note: str,
    user_id: uuid.UUID,
) -> ApplicationStageHistory:
    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")

    entry = ApplicationStageHistory(
        application_id=application_id,
        from_stage=app.stage,
        to_stage="_note",
        notes=note,
        changed_by=user_id,
    )
    db.add(entry)
    await db.commit()
    await db.refresh(entry)
    return entry


async def get_timeline(
    db: AsyncSession,
    application_id: uuid.UUID,
) -> list[ApplicationStageHistory]:
    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")

    rows = (await db.execute(
        select(ApplicationStageHistory)
        .where(ApplicationStageHistory.application_id == application_id)
        .order_by(ApplicationStageHistory.created_at.asc())
    )).scalars().all()
    return list(rows)


# Candidate may self-edit only while the application is early in the pipeline.
# HR/admins can edit throughout (see update_application).
CANDIDATE_EDITABLE_STAGES = ("applied", "screening")
_HR_EDIT_ROLES = ("hr_manager", "admin", "super_admin")
# Fields on ApplicationUpdate that belong to CandidateProfile, not the applications table.
_PROFILE_UPDATE_FIELDS = (
    "current_location", "total_experience", "current_company",
    "current_designation", "education", "skills",
)


async def update_application(
    db: AsyncSession,
    application_id: uuid.UUID,
    current_user,
    data,
) -> Application:
    from app.models.candidate_profile import CandidateProfile
    from app.models.user import User as UserModel

    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")

    is_hr = getattr(current_user, "role", None) in _HR_EDIT_ROLES
    if not is_hr:
        # Candidate self-service: must own it and only while early-stage.
        if str(app.applicant_id) != str(current_user.id):
            raise HTTPException(403, "Not your application")
        if app.stage not in CANDIDATE_EDITABLE_STAGES:
            raise HTTPException(
                400,
                "You can only edit your details while the application is in the "
                "Applied or Screening stage. Please contact the recruiter for later changes.",
            )

    payload = data.model_dump(exclude_unset=True)

    _validate_free_text_lengths(**{
        field: payload[field]
        for field in _FREE_TEXT_LABELS
        if field in payload
    })

    # Route profile fields to CandidateProfile and DOB to the user record.
    profile_updates = {k: payload.pop(k) for k in _PROFILE_UPDATE_FIELDS if k in payload}
    dob = payload.pop("date_of_birth", None)

    if profile_updates:
        profile = await db.get(CandidateProfile, app.applicant_id)
        if not profile:
            profile = CandidateProfile(user_id=app.applicant_id)
            db.add(profile)
        for field, val in profile_updates.items():
            if val is not None:
                setattr(profile, field, val)

    if dob is not None:
        user = await db.get(UserModel, app.applicant_id)
        if user:
            user.date_of_birth = dob

    # Remaining keys are application-level columns.
    for field, val in payload.items():
        if val is not None:
            setattr(app, field, val)

    app.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(app)
    return app


async def withdraw_application(
    db: AsyncSession,
    application_id: uuid.UUID,
    applicant_id: uuid.UUID,
) -> None:
    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")
    if app.applicant_id != applicant_id:
        raise HTTPException(403, "Not your application")
    if app.stage in ("hired", "rejected"):
        raise HTTPException(400, "Cannot withdraw a closed application")

    app.stage = "withdrawn"
    app.stage_updated_at = datetime.utcnow()
    await db.commit()


# ── Resume revisions ──────────────────────────────────────────────────────────

# Once an application is closed there is nothing left to review, so the
# candidate can no longer swap their resume. HR keeps access throughout — they
# may still need to correct a bad file on a hired candidate's record.
_RESUME_CLOSED_STAGES = ("hired", "rejected", "withdrawn", "interview_drop", "offer_drop")


async def _resume_access(db: AsyncSession, application_id: uuid.UUID, current_user):
    """Load the application and classify the caller as HR, owner, or neither."""
    app = await db.get(Application, application_id)
    if not app:
        raise HTTPException(404, "Application not found")

    role = getattr(current_user, "role", None)
    is_hr = role in _HR_EDIT_ROLES
    is_interviewer = role == "interviewer"
    is_owner = str(app.applicant_id) == str(current_user.id)
    return app, is_hr, is_interviewer, is_owner


async def list_application_resumes(
    db: AsyncSession,
    application_id: uuid.UUID,
    current_user,
) -> list[dict]:
    from app.models.application_resume import ApplicationResume
    from app.models.user import User as UserModel
    from app.services.storage_service import refresh_url

    app, is_hr, is_interviewer, is_owner = await _resume_access(db, application_id, current_user)
    if not (is_hr or is_interviewer or is_owner):
        raise HTTPException(403, "Not your application")

    rows = (await db.execute(
        select(ApplicationResume, UserModel.full_name)
        .outerjoin(UserModel, UserModel.id == ApplicationResume.uploaded_by)
        .where(ApplicationResume.application_id == application_id)
        .order_by(ApplicationResume.version.desc())
    )).all()

    return [
        {
            "id": r.id,
            "application_id": r.application_id,
            "version": r.version,
            # Same re-signing rule as every other stored blob URL — the SAS baked
            # in at upload expires after 7 days (see storage_service.refresh_url).
            "file_url": refresh_url(r.file_url),
            "file_name": r.file_name,
            "note": r.note,
            "uploaded_by": r.uploaded_by,
            "uploaded_by_name": name,
            "uploaded_by_role": r.uploaded_by_role,
            "is_current": r.file_url == app.resume_url,
            "created_at": r.created_at,
        }
        for r, name in rows
    ]


async def add_application_resume(
    db: AsyncSession,
    application_id: uuid.UUID,
    current_user,
    file,
    note: Optional[str] = None,
) -> dict:
    """Store a new resume revision and point the application at it."""
    from app.models.application_resume import ApplicationResume
    from app.services.storage_service import upload_resume

    app, is_hr, _is_interviewer, is_owner = await _resume_access(db, application_id, current_user)

    if not is_hr:
        if not is_owner:
            raise HTTPException(403, "Not your application")
        if app.stage in _RESUME_CLOSED_STAGES:
            raise HTTPException(
                400,
                "This application is closed — you can no longer update your resume. "
                "Please contact the recruiter.",
            )

    file_url = await upload_resume(file, str(app.applicant_id))

    # Versions are contiguous per application; MAX+1 rather than a count so a
    # deleted row could never hand out a number that is already taken.
    next_version = ((await db.execute(
        select(func.max(ApplicationResume.version))
        .where(ApplicationResume.application_id == application_id)
    )).scalar() or 0) + 1

    revision = ApplicationResume(
        application_id=application_id,
        version=next_version,
        file_url=file_url,
        file_name=getattr(file, "filename", None),
        note=note or None,
        uploaded_by=current_user.id,
        uploaded_by_role=getattr(current_user, "role", None),
    )
    db.add(revision)

    app.resume_url = file_url
    app.updated_at = datetime.utcnow()

    # Visible on the candidate timeline HR already reads, so a swapped resume
    # isn't something you only discover by opening the resume tab.
    who = "the candidate" if is_owner and not is_hr else getattr(current_user, "full_name", "HR")
    db.add(ApplicationStageHistory(
        application_id=application_id,
        from_stage=app.stage,
        to_stage=app.stage,
        changed_by=current_user.id,
        notes=f"Resume updated to v{next_version} by {who}" + (f" — {note}" if note else ""),
    ))

    await db.commit()
    await db.refresh(revision)

    return {
        "id": revision.id,
        "application_id": revision.application_id,
        "version": revision.version,
        "file_url": file_url,
        "file_name": revision.file_name,
        "note": revision.note,
        "uploaded_by": revision.uploaded_by,
        "uploaded_by_name": getattr(current_user, "full_name", None),
        "uploaded_by_role": revision.uploaded_by_role,
        "is_current": True,
        "created_at": revision.created_at,
    }
